from django.shortcuts import render
from django.http import JsonResponse
import json
from django.template.context_processors import request
from ..models import Asset
from ..models import Status
from ..models import Customer
from ..models import Custom
from ..models import Regime
from ..models import AssetImage
from ..models import AssetFile
from ..models import Activity
from assets.models import Activity, AssetFile, AssetQR
from django.db.models import Max
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from io import BytesIO
from typing import List
from django.db.models import Q
from django.contrib.auth.models import User
from django_renderpdf.views import PDFView


def testJson(request):
    data = {
            'message': 'Hello from Django!',
            'status': 'success',
            'items': [
                {'id': 1, 'name': 'Item A'},
                {'id': 2, 'name': 'Item B'}
            ]
        }
    return JsonResponse(data)


def search(asset_identifier = None, asset = None, customer_id = None, description = None, pediment = None, invoice = None, status = None):
    if(status is None):
        status = Status.objects.get(name = 'Entrada')
    else:
        status = Status.objects.get(name = status)

    assets_query = Asset.objects.filter(status = status)
    if (customer_id is not None or customer_id == 0):
        customer = Customer.objects.get(pk=customer_id)
        assets_query = Asset.objects.filter(customer = customer)
    
    # Add a lot of OR condittions
    query = Q()
    if (asset_identifier is not None):
        query |= Q(asset_identifier__contains = asset_identifier)
    if (asset is not None):
        query |= Q(asset__contains = asset)
    if (description is not None):
        query |= Q(description__contains = description)
    if (pediment is not None):
        query |= Q(pediment__contains = pediment)
    if (invoice is not None):
        query |= Q(invoice__contains = invoice)

    assets_query = assets_query.filter(query).order_by('-pk')

    return assets_query




def createOrUpdate(request):
    action = 'Update Asset'
    data = request.POST
    asset = Asset.objects.filter(pk=int(data.get('asset_id') or 0)).first()

    if asset is None:
         asset = Asset()
         action = 'Create Asset'

    asset_name = data.get('asset')
    customer = Customer.objects.filter(pk=int(data.get('customer') or 0)).first()
    description = data.get('description')
    brand = data.get('brand')
    model = data.get('model')
    serial = data.get('serial')
    location = data.get('location')
    pediment = data.get('pediment')
    invoice = data.get('invoice')
    invDate = data.get('invDate')
    motivo_salida = data.get('motivo_salida')
    
    
    regime = Regime.objects.filter(pk=int(data.get('regime') or 0)).first()
    custom = Custom.objects.filter(pk=int(data.get('custom') or 0)).first()
    patent = data.get('patent')
    notes = data.get('notes')
    status = data.get('status')
    status = Status.objects.filter(name=status).first()

    # asset.asset = asset_name #<- ahora podemos usar la funcion setField igual de valida pero nos ayudara a crear una registro de la transaccion mas detallado
    asset.setField(request.user, 'asset', asset_name)
    asset.customer = customer
    asset.setField(request.user, 'description', description)
    asset.setField(request.user, 'brand', brand)
    asset.setField(request.user, 'model', model)
    asset.setField(request.user, 'serial_number', serial)
    asset.setField(request.user, 'storage_location', location)
    asset.setField(request.user, 'pediment', pediment)
    asset.setField(request.user, 'invoice', invoice)
    asset.setField(request.user, 'pediment_invoice_date', invDate)
    asset.setField(request.user, 'regime', regime)
    asset.setField(request.user, 'custom', custom)
    asset.setField(request.user, 'patent', patent)
    asset.setField(request.user, 'status', status)
    asset.setField(request.user, 'notes', notes)
    asset.setField(request.user, 'motivo_salida', motivo_salida)

    # assignar un asset_identifier, este id es unico para cada asset-cliente
    last_asset = Asset.objects.filter(customer = customer).order_by('-asset_identifier').first()
    asset_identifier = 0
    if(last_asset is not None):
        asset_identifier = last_asset.asset_identifier
    asset.asset_identifier = int(asset_identifier) + 1

    asset.save()

    # guardar imagenes y archivos
    images =  request.FILES.getlist('images[]')
    for f in images:
        AssetImage.objects.create(asset=asset, image=f)

    files =  request.FILES.getlist('files[]')
    for f in files:
        AssetFile.objects.create(asset=asset, file=f)

    data = {
            'message': 'Registrado!',
            'status': 'success',
        }
    return JsonResponse(data)

def getContext(user: User, asset: Asset = None):
    customers = Customer.getCustomersForUser(user)
    # customers = Customer.objects.all() <- esto regresa todos los usuarios, pero vamos a filtrarlos
    customs = Custom.objects.all()
    regimes = Regime.objects.all()

    title = ('Crear' if asset is None  else 'Editar') + ' Activo'
    context = {
        'title': title,
        'asset': asset,
        'customers': customers,
        'regimes': regimes,
        'customs': customs
    }
    return context


def getIndexContext(assets, user: User):
    customers = Customer.getCustomersForUser(user)

    title = 'Busqueda'
    context = {
        'title': title,
        'assets': assets,
        'customers': customers,
    }
    return context

def deleteImage(request, img_id):
    img = AssetImage.objects.get(pk=int(img_id))
    msg = ''
    if(img is None):
        msg = 'El ID de la imagen a eliminar es invalido'
    else:
        img.delete()
        msg = 'Eliminada!'
    
    data = {
            'message': msg,
            'status': 'success',
        }
    return JsonResponse(data)

def deleteFile(request, file_id):
    img = AssetFile.objects.get(pk=int(file_id))
    msg = ''
    if(img is None):
        msg = 'El ID de la imagen a eliminar es invalido'
    else:
        img.delete()
        msg = 'Eliminada!'
    
    data = {
            'message': msg,
            'status': 'success',
        }
    return JsonResponse(data)

def registrarSalida():
    return None

def exportAssets(assets: List[Asset]):
    # "Create a new workbook"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "My Data"

    # Add some data to the worksheet
    ws.append(["Aduana(3)*",
               "Patente(4)*",
               "Pedimento(7)*",
               "Fecha(dd/mm/yyyy)*",
               "Proveedor (150)",
               "Clave(2)*",
               "Marca(150)",
               "Modelo(150)",
               "Número de identificación(150)",
               "Ubicación(150)",
               "Control Fiscal(150)",
               "Observaciones",
               "Control Admon(150)",
               "Valor(150)",
               "Descripción(150)",
               "Valor2(150)",
               "Origen(150)",
               "Orden de Compra(150)",
               "No. de Contrato(150)",
               "Documento Equivalente(150)",
               "Proyecto(150)",
               "Estatus(150)",
               "Documento Baja(150)",
               "Fecha Baja(dd//mm/yyyy)",
               "Ubicación Temporal(150)",
               "Fecha Traslado(dd//mm/yyyy)"])
    for asset in assets:
        ws.append([asset.custom.name,
                   asset.patent,
                   asset.pediment,
                   asset.pediment_invoice_date.strftime("%d-%m-%Y"),
                   '',
                   '',
                   asset.brand,
                   asset.model,
                   asset.serial_number,
                   asset.asset,
                   asset.storage_location,
                   asset.notes,
                   asset.asset_identifier,
                   '',
                   asset.description,
                   '',
                   '',
                   '',
                   '',
                   '',
                   '',
                   asset.status.name,
                   '',
                   '',
                   '',
                   ''
                   ])
    # Apply some style
    fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    for row_of_cells in ws['A1':'Z1']:
        for cell in row_of_cells:
            cell.font = Font(bold=True)
            cell.fill = fill
    

    # Save the workbook to a BytesIO object
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)  # Rewind the buffer to the beginning

    # Create an HttpResponse with the Excel file
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="export.xlsx"'
    return response



# def loadQRToEnt(asset: Asset, url):
    

def invalidRoute(request):
    return
